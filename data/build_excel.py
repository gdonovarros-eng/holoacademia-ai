import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────
JSON_PATH  = "/Users/m2/Documents/New project/data/biomagnetic_pairs_db.json"
XLSX_PATH  = "/Users/m2/Documents/New project/data/pares_biomagneticos.xlsx"

# ── palette ────────────────────────────────────────────────────────────────
HDR_BG    = "1D1D1F"
HDR_FG    = "FFFFFF"
ROW_ODD   = "FFFFFF"
ROW_EVEN  = "F5F5F7"

# ── column widths ──────────────────────────────────────────────────────────
COL_WIDTHS = {
    "A": 14,   # Región
    "B": 18,   # Zona
    "C": 22,   # Bloque
    "D": 5,    # #
    "E": 35,   # Polo Negro
    "F": 35,   # Polo Rojo
}

FONT_NAME = "Calibri"
FONT_SIZE = 10

def make_font(bold=False, color="000000"):
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, color=color)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── load data ──────────────────────────────────────────────────────────────
with open(JSON_PATH, encoding="utf-8") as f:
    db = json.load(f)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Pares
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Pares Biomagneticos"

# header
HEADERS = ["Región", "Zona", "Bloque", "#", "Polo Negro", "Polo Rojo"]
ws.append(HEADERS)
hdr_row = ws[1]
for cell in hdr_row:
    cell.font      = make_font(bold=True, color=HDR_FG)
    cell.fill      = make_fill(HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 20

# freeze header
ws.freeze_panes = "A2"

# set column widths
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# data rows
total_pairs = 0
summary = []   # (region_name, set_of_zonas, set_of_bloques, count)

for region in db["regiones"]:
    region_name = region["nombre"]
    zonas_set   = set()
    bloques_set = set()
    region_count = 0

    for zona in region["zonas"]:
        zona_name = zona["nombre"]
        zonas_set.add(zona_name)

        for bloque in zona["bloques"]:
            bloque_name = bloque["nombre"]
            bloques_set.add(bloque_name)

            for pair_num, pair_str in enumerate(bloque["pares"], start=1):
                sep = " – "   # space – en-dash – space
                if sep in pair_str:
                    polo_neg, polo_rojo = pair_str.split(sep, 1)
                else:
                    # try em-dash fallback
                    sep2 = " — "
                    if sep2 in pair_str:
                        polo_neg, polo_rojo = pair_str.split(sep2, 1)
                    else:
                        polo_neg  = pair_str
                        polo_rojo = ""

                ws.append([region_name, zona_name, bloque_name,
                            pair_num, polo_neg, polo_rojo])
                total_pairs += 1
                region_count += 1

                # alternating row colours (row 2 = first data row)
                row_idx   = total_pairs + 1          # +1 because header is row 1
                fill_hex  = ROW_ODD if (total_pairs % 2 == 1) else ROW_EVEN
                row_fill  = make_fill(fill_hex)
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font      = make_font()
                    cell.fill      = row_fill
                    cell.alignment = Alignment(vertical="center",
                                               wrap_text=False)

    summary.append((region_name,
                    len(zonas_set),
                    len(bloques_set),
                    region_count))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Resumen
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen")

SUM_HEADERS = ["Región", "Zonas", "Bloques", "Total Pares"]
ws2.append(SUM_HEADERS)
hdr2 = ws2[1]
for cell in hdr2:
    cell.font      = make_font(bold=True, color=HDR_FG)
    cell.fill      = make_fill(HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 20
ws2.freeze_panes = "A2"

# column widths for summary
for col_letter, width in [("A", 18), ("B", 10), ("C", 12), ("D", 14)]:
    ws2.column_dimensions[col_letter].width = width

tot_zonas = tot_bloques = tot_pares = 0

for i, (reg, zonas, bloques, pares) in enumerate(summary, start=1):
    ws2.append([reg, zonas, bloques, pares])
    fill_hex = ROW_ODD if (i % 2 == 1) else ROW_EVEN
    row_fill = make_fill(fill_hex)
    for col_idx in range(1, 5):
        cell = ws2.cell(row=i + 1, column=col_idx)
        cell.font      = make_font()
        cell.fill      = row_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                    vertical="center")
    tot_zonas   += zonas
    tot_bloques += bloques
    tot_pares   += pares

# TOTAL row
total_row_idx = len(summary) + 2
ws2.append(["TOTAL", tot_zonas, tot_bloques, tot_pares])
for col_idx in range(1, 5):
    cell = ws2.cell(row=total_row_idx, column=col_idx)
    cell.font      = make_font(bold=True, color=HDR_FG)
    cell.fill      = make_fill(HDR_BG)
    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                vertical="center")

# ── save ───────────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"Total pares escritos: {total_pairs}")
print(f"Archivo guardado en: {XLSX_PATH}")
